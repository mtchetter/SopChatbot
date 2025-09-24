"""| 5/21/2025 | BandaKi | Working | Email sends the excel that contains the open tickets and this cleans it an upsets into SQL."""

import openpyxl
import os
import pyodbc
import pandas as pd
from datetime import datetime
import getpass

def extract_hyperlinks_and_merged_values_and_insert_to_sql():
    print("Step/1: Setting up file path and loading workbook")
    file_path = r"C:\Users\mtschetter\OneDrive - heihotels.com\Projects\operation_frozen_heart\system_tickets\ticket_data.xlsx"
    print(f"Loading file from: {file_path}")
    
    wb = openpyxl.load_workbook(file_path)
    print(f"Workbook opened successfully")
    
    print("Step/2: Selecting active worksheet")
    ws = wb.active
    print(f"Active worksheet is: {ws.title}")
    
    print("Step/3: Extracting hyperlinks from column G and placing in column I")
    print("Step/4: Handling merged cells and placing values in column J")
    
    current_merged_value = None
    
    for row in range(1, ws.max_row + 1):
        # Extract hyperlinks from column G to column I
        cell_g = ws.cell(row=row, column=7)  # Column G is the 7th column
        
        print(f"Processing row {row}, cell G{row}")
        
        # Check if cell has hyperlink
        if cell_g.hyperlink:
            hyperlink = cell_g.hyperlink.target
            print(f"Found hyperlink in G{row}: {hyperlink}")
            
            # Place hyperlink in column I (9th column)
            ws.cell(row=row, column=9).value = hyperlink
            print(f"Placed hyperlink in I{row}")
        else:
            print(f"No hyperlink found in G{row}")
        
        # Check for merged cells
        print(f"Checking for merged cells in row {row}")
        
        # Look for merged ranges that include cells in this row
        found_merged_value = False
        for merged_range in ws.merged_cells.ranges:
            # Check if any cell in this row is part of a merged range
            for col in range(1, ws.max_column + 1):
                cell_position = f"{openpyxl.utils.get_column_letter(col)}{row}"
                if cell_position in merged_range:
                    # Get the value from the top-left cell of the merged range
                    top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    
                    if top_left.value:
                        print(f"Found merged cell value in row {row}: {top_left.value}")
                        current_merged_value = top_left.value
                        found_merged_value = True
                        break
            
            if found_merged_value:
                break
        
        # Place the current merged value in column J (10th column)
        if current_merged_value:
            ws.cell(row=row, column=10).value = current_merged_value
            print(f"Placed merged value in J{row}: {current_merged_value}")
    
    print("Step/5: Copying data from columns A-J starting from row 11 to next tab with headers")
    
    # Check if there's already a second sheet, if not create one
    if len(wb.sheetnames) < 2:
        print("Creating new worksheet for data")
        ws_target = wb.create_sheet(title="Copied_Data")
    else:
        print(f"Using existing worksheet: {wb.sheetnames[1]}")
        ws_target = wb[wb.sheetnames[1]]
    
    # Clear any existing data in the target worksheet
    print("Clearing existing data in target worksheet")
    for row in ws_target.rows:
        for cell in row:
            cell.value = None
    
    # Add column headers in row 1
    print("Step/5.1: Adding column headers to row 1")
    headers = ["request_id", "created_time", "last_updated_time", "request_status", 
               "requester", "department", "subject", "sub_category", "ticket_link", "assigned_person"]
    
    for col_idx, header in enumerate(headers, start=1):
        ws_target.cell(row=1, column=col_idx).value = header
        print(f"Added header '{header}' to column {openpyxl.utils.get_column_letter(col_idx)}")
    
    # Track which rows to skip (rows that are part of merged cells)
    merged_rows = set()
    
    print("Identifying merged rows to exclude")
    for merged_range in ws.merged_cells.ranges:
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            if r >= 11:  # Only care about merged rows from row 11 onward
                merged_rows.add(r)
                print(f"Row {r} is part of a merged range - will be excluded")
    
    # Copy data from columns A-J starting from row 11, excluding merged rows
    # Now starting at row 2 in target worksheet (below headers)
    print("Copying data to target worksheet starting at row 2 (below headers)")
    target_row = 2  # Start at row 2 in the target worksheet (below headers)
    
    for source_row in range(11, ws.max_row + 1):
        if source_row in merged_rows:
            print(f"Skipping row {source_row} as it's part of a merged range")
            continue
        
        print(f"Copying row {source_row} to target row {target_row}")
        for col in range(1, 11):  # Columns A-J (1-10)
            source_cell = ws.cell(row=source_row, column=col)
            ws_target.cell(row=target_row, column=col).value = source_cell.value
        
        target_row += 1
    
    print("Step/6: Saving the updated workbook")
    output_path = os.path.join(os.path.dirname(file_path), "ticket_data_with_hyperlinks_and_merged.xlsx")
    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")
    
    print("Step/7: Convert Excel data to pandas DataFrame")
    # Create a list to store the data from the target worksheet
    data = []
    
    # Skip header row (row 1) and collect all data rows
    for row in range(2, ws_target.max_row + 1):
        row_data = []
        for col in range(1, 11):  # Columns A-J (1-10)
            cell_value = ws_target.cell(row=row, column=col).value
            row_data.append(cell_value)
        
        if any(row_data):  # Only add rows that have at least one non-empty value
            data.append(row_data)
    
    # Create DataFrame with the column headers
    df = pd.DataFrame(data, columns=headers)
    print(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
    
    print("Step/8: Format date columns")
    # Convert date columns to proper format for SQL Server
    # Format: MM-DD-YYYY HH as specified in your table definition
    for date_col in ['created_time', 'last_updated_time']:
        print(f"Formatting {date_col} column")
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        # Format as MM-DD-YYYY HH
        df[date_col] = df[date_col].dt.strftime('%m-%d-%Y %H')
    
    print("Step/9: Get current user and timestamp")
    current_user = getpass.getuser()
    print(f"Current user: {current_user}")
    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"Current datetime: {current_datetime}")
    
    print("Step/10: Establish SQL Server connection")
    # Use the 'SQL Server' driver which we know works
    conn_str = (
        'DRIVER={SQL Server};'
        'SERVER=colofinsql02;'
        'DATABASE=FinancialModeling;'
        'Trusted_Connection=yes;'
    )
    
    print("Step/11: Insert data into SQL Server table")
    try:
        print("Connecting with SQL Server driver...")
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        print("Connection established successfully")
        
        # Prepare insert statement
        table_name = '[FinancialModeling].[hr].[hr_tickets]'
        print(f"Preparing to insert data into {table_name}")
        
        # Generate parameterized insert statement with the new columns
        insert_sql = f"""
        INSERT INTO {table_name} 
        (request_id, created_time, last_updated_time, request_status, 
        requester, department, subject, sub_category, ticket_link, assigned_person,
        inserted_datetime, inserted_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        
        # Insert each row
        rows_inserted = 0
        for index, row in df.iterrows():
            try:
                print(f"Inserting row {index+1} of {len(df)}")
                values = tuple(row) + (current_datetime, current_user)
                cursor.execute(insert_sql, values)
                rows_inserted += 1
                
                # Commit every 50 rows to avoid large transactions
                if rows_inserted % 50 == 0:
                    conn.commit()
                    print(f"Committed {rows_inserted} rows so far")
            except pyodbc.Error as row_error:
                print(f"Error inserting row {index+1}: {row_error}")
                print(f"Problematic data: {values}")
                # Continue with next row despite error
                continue
        
        # Final commit for remaining rows
        conn.commit()
        print(f"Successfully inserted {rows_inserted} rows into {table_name}")
        
        # Close connection
        cursor.close()
        conn.close()
        print("Database connection closed")
        
    except pyodbc.Error as e:
        print(f"Error with SQL Server connection: {e}")
    
    print("Step/12: Process completed successfully!")

    print("Step/13: Archiving original Excel file and cleaning up source directory")

    archive_dir = r"C:\Users\mtschetter\OneDrive - heihotels.com\Projects\operation_frozen_heart\system_tickets\archive"
    original_file = r"C:\Users\mtschetter\OneDrive - heihotels.com\Projects\operation_frozen_heart\system_tickets\ticket_data.xlsx"
    processed_file = os.path.join(os.path.dirname(original_file), "ticket_data_with_hyperlinks_and_merged.xlsx")

    try:
        # Ensure archive directory exists
        os.makedirs(archive_dir, exist_ok=True)

        # Generate timestamp in mm_dd_yy_HH format
        timestamp = datetime.now().strftime('%m_%d_%y_%H')

        # Split original filename and extension
        base_name, ext = os.path.splitext(os.path.basename(original_file))

        # Create new filename with timestamp
        archived_filename = f"{base_name}_{timestamp}{ext}"
        archived_file = os.path.join(archive_dir, archived_filename)

        print(f"Copying original file to archive as: {archived_file}")
        import shutil
        shutil.copy2(original_file, archived_file)

        # Delete both files from the original directory
        print("Deleting original and processed files from source directory")
        os.remove(original_file)
        os.remove(processed_file)

        print("Archiving and cleanup complete.")
    except Exception as cleanup_error:
        print(f"Error during archiving or cleanup: {cleanup_error}")



extract_hyperlinks_and_merged_values_and_insert_to_sql()